"""
Daily Outstanding Shipment Processor (SOMS / Packing List / UPS prep)
======================================================================
Encodes the SOP "001-SOP_Upload_Daily_Shipments_WAWI_PL_UPS" (WAWI is
the German term for what's referred to throughout this script as SOMS
- a generic Sales Order Management System) plus the clarified business
rules agreed through 2026-08-26. Runs entirely locally - no internet,
no AI calls, no cost.

We now work exclusively in real .xlsx format (not CSV), which is what
lets us do genuine Excel cell formatting (see ContactNoHome/Mobile below)
instead of CSV-only text workarounds.

USAGE
-----
    python OutstandingShipments.py

By default it looks in your Downloads folder for a file matching
    OS_*.xlsx
(whatever date is in the name - the script does not check that it
matches today's date, per instruction). You can also pass a path
explicitly:

    python OutstandingShipments.py "C:\\Users\\you\\Downloads\\OS_2026-08-27.xlsx"

OUTPUT
------
    <same filename as source>_for_soms.xlsx
saved to the same Downloads folder (e.g. OS_2026-08-27.xlsx ->
OS_2026-08-27_for_soms.xlsx). Adds one new column, "Notes and
Warnings" (the last column, after "Nr. of packages"), flagging anything
edited or anything that needs a human look.

WHAT IT DOES (2026-08-26 rules)
--------------------------------
- Combines OrderNo values (space-separated) into OrderNo, and ReceiptNo
  values (space-separated) into ReceiptNo, for any IRID with 2-3 orders.
- ZIP codes under 5 digits for Germany/Italy/France/Spain get padded
  with leading zeros (e.g. 7545 -> 07545), written as real text so the
  leading zero displays correctly.
- Address lines over 35 characters (UPS limit) get wrapped across
  ADDR1/ADDR2/ADDR3 at a word boundary - only when the target field is
  empty; if it's already occupied, or the line can't be split cleanly,
  it's left untouched and flagged instead of guessing.
- In ADDR1/ADDR2/ADDR3 only, a space is inserted wherever a letter is
  immediately followed by a digit with no space (e.g. "FLAT16" ->
  "FLAT 16"). Does NOT touch digit-then-letter (e.g. "2a" - a normal
  European house-number-suffix convention) or any other column.
- If ADDR2 or ADDR3's entire content exactly duplicates ADDR1, ZIPCODE,
  CITY, or CountryName, it's cleared out (before the length-split, so a
  duplicate ADDR2 doesn't block the split from using that field).
- If ADDR1 needs splitting but ADDR2 already has (non-duplicate)
  content, ADDR2 is relocated to ADDR3 to make room, then the split
  proceeds - reverted automatically if that still doesn't leave enough
  room to fit everything.
- Any row where the CITY name or ZIP code literally appears inside the
  address lines gets flagged "Address verification required".
- ProductPrice sometimes comes out corrupted (e.g. Excel stored it as
  an actual date, or it reads as date-like text such as "Dez 71").
  Where that happens, it's recalculated as the net price before VAT:
  (TotalAmount - ShipFee) / (1 + country's standard VAT rate), and
  logged in Notes and Warnings. Valid prices are left untouched.
- ContactNoHome and ContactNoMobile are written as real numbers with an
  Excel custom number format of "+0" (Format Cells > Custom > +0) - so
  Excel displays e.g. +491234567890 while the underlying stored value
  is just the plain digits, exactly like manually formatting the cells.
- Rows are sorted by IRID, A-Z.
- Every other column is written as plain text, matching what's read in
  (so things like ProdCode/Quantity round-trip unchanged).

IMPORTANT - this script does NOT guess. When it hits a case the rules
don't cover (bad ZIP, unknown product needing substitution, more than
3 shipments in one IR ID group, an address it can't safely split, etc.)
it either STOPS the whole run (for filewide issues) or flags that row
in Notes and Warnings (for row-level address issues) - it never mixes
up or invents customer data.

Nothing is uploaded anywhere and no source files are deleted.
"""

import sys
import re
from pathlib import Path

import pandas as pd
import openpyxl

# ----------------------------------------------------------------------
# CONFIG - edit this section if business rules change. No other code
# should need to change for typical rule tweaks.
# ----------------------------------------------------------------------

DOWNLOADS_DIR = Path.home() / "Downloads"

EXPECTED_COLUMNS = [
    "OrderNo", "ReceiptNo", "IRID", "ProdCode", "ComboCode", "ProdName",
    "Quantity", "IRNAME", "IRNAME2", "ADDR1", "ADDR2", "ADDR3", "ZIPCODE",
    "CITY", "CountryName", "ContactNoHome", "ContactNoMobile", "EMAIL",
    "ProductPrice", "ShipFee", "TotalAmount",
]

# Columns written as real numbers with a custom Excel format (Format
# Cells > Custom > +0), instead of plain text like everything else.
PHONE_COLUMNS = ["ContactNoHome", "ContactNoMobile"]
PHONE_NUMBER_FORMAT = "+0"

# Max shipments that can be combined under one IR ID (UPS label limit).
MAX_COMBINED_SHIPMENTS = 3

# Product name substitutions: out-of-stock product -> product actually
# shipped. Add rows here as needed. Matched against ProdName exactly.
# NOTE: "AirPure Zayn" substitution target was never specified in the
# SOP - add it here once known, otherwise the script will flag and stop
# if that product ever appears.
PRODUCT_SUBSTITUTIONS = {
    # "Old Product Name": {"ProdName": "New Product Name", "ProdCode": "0000000000"},
}

# Products that trigger "Nr. of packages" review.
PACKAGE_TRIGGER_ZAYN = "HomePure Zayn"
PACKAGE_TRIGGER_NOVA = "HomePure Nova"

# Known-bad ZIP placeholders that require the user to supply a real one.
BAD_ZIPS = {"", "0000", "00000", "nan", "none"}

# Quick fixes for today's run: IRID -> correct ZIP. Add an entry here when
# the script stops and tells you an IRID has a bad ZIP, then re-run.
# Clear this dict out periodically once orders have shipped.
ZIP_OVERRIDES = {
    # "AB123456": "12345",
    # "CD987654": "54321",
}

# Countries whose postal codes should be 5 digits - pad with a leading
# zero if shorter. Matched case-insensitively against CountryName.
ZIP_PAD_COUNTRIES = {"germany", "italy", "france", "spain"}
ZIP_PAD_LENGTH = 5

# Carrier max characters per address line (UPS - this SOP is UPS-focused).
ADDRESS_LINE_LIMIT = 35

# Standard VAT rates by country (checked against current sources,
# 2026). Used only to recompute ProductPrice when it comes out corrupted
# (e.g. read as a date) - NOT applied to otherwise-valid prices.
VAT_RATES = {
    "germany": 0.19,
    "france": 0.20,
    "spain": 0.21,
    "italy": 0.22,
    "ireland": 0.23,
    "poland": 0.23,
    "sweden": 0.25,
    "united kingdom": 0.20,
}

# ----------------------------------------------------------------------


def find_input_file(explicit_path: str | None) -> Path:
    if explicit_path:
        p = Path(explicit_path)
        if not p.exists():
            sys.exit(f"ERROR: file not found: {p}")
        return p

    # Source filename pattern: "OS_<date>.xlsx". We do NOT gate on the
    # date matching today - just find the file.
    matches = sorted(DOWNLOADS_DIR.glob("OS_*.xlsx"))

    if len(matches) == 1:
        return matches[0]

    if len(matches) == 0:
        print("\n" + "=" * 70)
        print("STOPPED: no file matching 'OS_*.xlsx' found in")
        print(f"  {DOWNLOADS_DIR}")
        print("Save the source file there, or pass its path explicitly:")
        print('  python OutstandingShipments.py "C:\\path\\to\\file.xlsx"')
        print("=" * 70)
        sys.exit(1)

    # More than one match - don't guess which is today's, that risks
    # processing the wrong day's customer orders.
    print("\n" + "=" * 70)
    print("STOPPED: found more than one matching source file in Downloads:")
    for m in matches:
        print(f"  {m.name}")
    print("Refusing to guess which one to use. Delete/move the old one(s),")
    print("or re-run with the correct path as an explicit argument:")
    print(f'  python OutstandingShipments.py "{matches[-1]}"')
    print("=" * 70)
    sys.exit(1)


def _cell_to_str(value) -> str:
    """Turn any raw cell value (str/int/float/datetime/None) into a plain
    string for internal processing. A corrupted ProductPrice cell that
    Excel stored as an actual date becomes its str() representation here
    (e.g. '2026-08-14 00:00:00'), which deliberately fails the later
    'is this a valid number' check - that's what flags it for repair."""
    if value is None:
        return ""
    return str(value)


def load_raw(path: Path) -> pd.DataFrame:
    """Source file is a real .xlsx workbook, single sheet."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        sys.exit(f"ERROR: could not open '{path.name}' as an Excel file: {e}")
    ws = wb.active

    header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    missing = set(EXPECTED_COLUMNS) - set(header)
    if missing:
        sys.exit(f"ERROR: source file is missing expected columns: {missing}. "
                  f"This may not be the correct file - stopping.")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({col: _cell_to_str(val) for col, val in zip(header, row)})

    df = pd.DataFrame(rows, columns=header)
    # Drop fully-blank trailing rows some exports leave behind.
    df = df[~(df[EXPECTED_COLUMNS].apply(lambda r: all(v == "" for v in r), axis=1))]
    return df.reset_index(drop=True)


def output_path_for(src: Path) -> Path:
    """Same filename as the source, with '_for_soms' appended before the
    extension (e.g. OS_2026-08-27.xlsx -> OS_2026-08-27_for_soms.xlsx) -
    whatever date/text is in the name, left untouched."""
    return DOWNLOADS_DIR / f"{src.stem}_for_soms{src.suffix}"


def write_output(df: pd.DataFrame, path: Path) -> None:
    """Write a real .xlsx file. Every column is plain text EXCEPT
    ContactNoHome/ContactNoMobile, which are written as real numbers
    with a custom '+0' number format (Format Cells > Custom > +0) -
    Excel displays them as +491234567890 while the stored value is just
    the digits, same as manually formatting the cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = path.stem[:31]  # Excel sheet name limit

    ws.append(list(df.columns))

    phone_col_idx = {col: i + 1 for i, col in enumerate(df.columns) if col in PHONE_COLUMNS}

    for row in df.itertuples(index=False):
        ws.append(list(row))

    for col_name, col_idx in phone_col_idx.items():
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            raw = str(cell.value).strip() if cell.value is not None else ""
            digits = re.sub(r"\D", "", raw)
            if digits:
                cell.value = int(digits)
                cell.number_format = PHONE_NUMBER_FORMAT
            else:
                cell.value = None

    wb.save(path)


def apply_zip_overrides(df: pd.DataFrame) -> pd.DataFrame:
    for irid, zipcode in ZIP_OVERRIDES.items():
        df.loc[df["IRID"] == irid, "ZIPCODE"] = zipcode
    return df


def append_notes(df: pd.DataFrame, notes: pd.Series) -> pd.DataFrame:
    """Append text to the 'Notes and Warnings' column, combining with
    whatever is already there via ' | '. Empty note = no-op for that row."""
    def combine(current, new):
        current = str(current).strip()
        new = str(new).strip()
        if not new:
            return current
        if not current:
            return new
        return current + " | " + new
    df["Notes and Warnings"] = [combine(c, n) for c, n in zip(df["Notes and Warnings"], notes)]
    return df


def pad_zip_codes(df: pd.DataFrame) -> pd.DataFrame:
    """For Germany/Italy/France/Spain, pad ZIP codes shorter than 5
    digits with leading zeros (written as real text, so the leading
    zero displays correctly - no apostrophe trick needed in real xlsx)."""
    def pad_one(zipcode, country):
        country_l = str(country).strip().lower()
        z = str(zipcode).strip()
        if country_l not in ZIP_PAD_COUNTRIES:
            return z, ""
        digits = re.sub(r"\D", "", z)
        if not digits or len(digits) >= ZIP_PAD_LENGTH:
            return z, ""
        padded = digits.zfill(ZIP_PAD_LENGTH)
        note = f"Address edited: ZIP padded {z} -> {padded}"
        return padded, note

    results = df.apply(lambda r: pad_one(r["ZIPCODE"], r["CountryName"]), axis=1)
    df["ZIPCODE"] = [t[0] for t in results]
    df = append_notes(df, pd.Series([t[1] for t in results], index=df.index))
    return df


def _split_long_line(text: str, limit: int) -> tuple[str, str] | None:
    """Split `text` into a first part <= limit chars and a remainder.
    Prefers splitting at a double-space if one exists within range (a
    natural boundary between address components seen in this data,
    e.g. street vs. floor/unit details) - falls back to the last single
    word-boundary at/before `limit` chars. Returns None if no safe split exists."""
    if len(text) <= limit:
        return text, ""
    window = text[: limit + 1]
    double_space_at = window.rfind("  ")
    if double_space_at > 0:
        first = text[:double_space_at].rstrip()
        remainder = text[double_space_at:].strip()
        return first, remainder
    split_at = window.rfind(" ")
    if split_at <= 0:
        return None  # no safe word boundary - single overlong word
    first = text[:split_at].rstrip()
    remainder = text[split_at:].strip()
    return first, remainder


def normalize_address_number_spacing(df: pd.DataFrame) -> pd.DataFrame:
    """Insert a space between a letter and an immediately-following digit
    in ADDR1/ADDR2/ADDR3 only (e.g. 'FLAT16' -> 'FLAT 16',
    'BARCELONA38' -> 'BARCELONA 38'), so a house/apartment/unit number
    isn't glued onto the preceding word. Deliberately does NOT touch the
    reverse case (digit immediately followed by a letter, e.g. '2a',
    '12b') - that's a normal European house-number-suffix convention and
    should stay as-is. Runs before the line-length split, since inserting
    spaces can push a line over the limit."""
    pattern = re.compile(r"(?<=[A-Za-zÀ-ÿ])(?=\d)")

    def fix_one(addr1, addr2, addr3):
        changes = []
        new_vals = []
        for label, val in (("ADDR1", addr1), ("ADDR2", addr2), ("ADDR3", addr3)):
            s = str(val)
            new_s = pattern.sub(" ", s)
            if new_s != s:
                changes.append(f"{label} '{s}' -> '{new_s}'")
            new_vals.append(new_s)
        note = f"Address edited: number separated from text ({', '.join(changes)})" if changes else ""
        return (*new_vals, note)

    results = df.apply(lambda r: fix_one(r["ADDR1"], r["ADDR2"], r["ADDR3"]), axis=1)
    df["ADDR1"] = [t[0] for t in results]
    df["ADDR2"] = [t[1] for t in results]
    df["ADDR3"] = [t[2] for t in results]
    df = append_notes(df, pd.Series([t[3] for t in results], index=df.index))
    return df


def remove_duplicate_address_fields(df: pd.DataFrame) -> pd.DataFrame:
    """If ADDR2 or ADDR3's entire content exactly duplicates ADDR1,
    ZIPCODE, CITY, or CountryName, clear it out. Runs before the
    line-length split, so a duplicate ADDR2 doesn't block the split from
    using that field (Point 1 in the 2026-08-28 instructions)."""
    def clean_one(addr1, addr2, addr3, zipcode, city, country):
        addr1_s = str(addr1).strip()
        addr2_s = str(addr2).strip()
        addr3_s = str(addr3).strip()
        reference = {v.strip().upper() for v in
                     (addr1_s, str(zipcode).strip(), str(city).strip(), str(country).strip()) if v.strip()}

        changes = []
        new_addr2, new_addr3 = addr2_s, addr3_s
        if addr2_s and addr2_s.upper() in reference:
            changes.append(f"ADDR2 '{addr2_s}' removed (duplicate)")
            new_addr2 = ""
        if addr3_s and addr3_s.upper() in reference:
            changes.append(f"ADDR3 '{addr3_s}' removed (duplicate)")
            new_addr3 = ""
        note = "Address edited: " + "; ".join(changes) if changes else ""
        return new_addr2, new_addr3, note

    results = df.apply(
        lambda r: clean_one(r["ADDR1"], r["ADDR2"], r["ADDR3"], r["ZIPCODE"], r["CITY"], r["CountryName"]),
        axis=1,
    )
    df["ADDR2"] = [t[0] for t in results]
    df["ADDR3"] = [t[1] for t in results]
    df = append_notes(df, pd.Series([t[2] for t in results], index=df.index))
    return df


def split_long_addresses(df: pd.DataFrame) -> pd.DataFrame:
    """If ADDR1 exceeds ADDRESS_LINE_LIMIT chars, wrap the overflow into
    ADDR2 (and ADDR3 if still needed), splitting only at word boundaries.
    If ADDR2 already has content, it's relocated to ADDR3 to make room
    (Point 2, 2026-08-28) - but only if that relocation actually lets the
    split succeed; if not, everything is reverted and the row is flagged
    for manual review instead of guessing."""
    def process(addr1, addr2, addr3):
        addr1 = str(addr1).strip()
        orig_addr2 = str(addr2).strip()
        orig_addr3 = str(addr3).strip()
        orig_len = len(addr1)

        if orig_len <= ADDRESS_LINE_LIMIT:
            return addr1, orig_addr2, orig_addr3, ""

        working_addr2 = orig_addr2
        working_addr3 = orig_addr3
        pre_notes = []

        if working_addr2:
            if working_addr3:
                return addr1, orig_addr2, orig_addr3, (
                    f"Address verification required: ADDR1 is {orig_len} chars "
                    f"(limit {ADDRESS_LINE_LIMIT}) but both ADDR2 and ADDR3 already have "
                    f"content - not auto-split"
                )
            # Tentatively relocate - only kept if the split actually succeeds below.
            working_addr3 = working_addr2
            working_addr2 = ""
            pre_notes.append(f"ADDR2 '{orig_addr2}' moved to ADDR3 to make room for ADDR1 split")

        split1 = _split_long_line(addr1, ADDRESS_LINE_LIMIT)
        if split1 is None:
            return addr1, orig_addr2, orig_addr3, (
                f"Address verification required: ADDR1 is {orig_len} chars "
                f"(limit {ADDRESS_LINE_LIMIT}) and has no safe word break to split on"
            )
        line1, remainder = split1

        if len(remainder) <= ADDRESS_LINE_LIMIT:
            notes = pre_notes + [f"ADDR1 ({orig_len} chars) split into ADDR1/ADDR2"]
            return line1, remainder, working_addr3, "Address edited: " + "; ".join(notes)

        # Remainder still too long - need a third line.
        if working_addr3:
            # Occupied either originally or by our tentative move - no room; revert.
            return addr1, orig_addr2, orig_addr3, (
                f"Address verification required: ADDR1 is {orig_len} chars "
                f"(limit {ADDRESS_LINE_LIMIT}) and the overflow doesn't fit in ADDR2 alone, "
                f"but ADDR3 already has content - not auto-split"
            )

        split2 = _split_long_line(remainder, ADDRESS_LINE_LIMIT)
        if split2 is None:
            return addr1, orig_addr2, orig_addr3, (
                f"Address verification required: ADDR1 is {orig_len} chars "
                f"(limit {ADDRESS_LINE_LIMIT}) and could not be cleanly split across 3 lines"
            )
        line2, line3 = split2
        if len(line3) > ADDRESS_LINE_LIMIT:
            return addr1, orig_addr2, orig_addr3, (
                f"Address verification required: ADDR1 is {orig_len} chars - "
                f"too long to fit across ADDR1/ADDR2/ADDR3 within the {ADDRESS_LINE_LIMIT}-char limit"
            )
        notes = pre_notes + [f"ADDR1 ({orig_len} chars) split across ADDR1/ADDR2/ADDR3"]
        return line1, line2, line3, "Address edited: " + "; ".join(notes)

    results = df.apply(lambda r: process(r["ADDR1"], r["ADDR2"], r["ADDR3"]), axis=1)
    df["ADDR1"] = [t[0] for t in results]
    df["ADDR2"] = [t[1] for t in results]
    df["ADDR3"] = [t[2] for t in results]
    df = append_notes(df, pd.Series([t[3] for t in results], index=df.index))
    return df


def flag_city_zip_in_address(df: pd.DataFrame) -> pd.DataFrame:
    """Flag rows where the CITY name or ZIP code literally appears inside
    the address lines - usually means the address wasn't split into
    fields properly upstream and needs a human look."""
    def check(addr1, addr2, addr3, city, zipcode):
        combined = f"{addr1} {addr2} {addr3}".upper()
        city_u = str(city).strip().upper()
        zip_digits = re.sub(r"\D", "", str(zipcode))
        if city_u and len(city_u) >= 3 and city_u in combined:
            return "Address verification required"
        if zip_digits and len(zip_digits) >= 4 and zip_digits in combined:
            return "Address verification required"
        return ""

    notes = df.apply(lambda r: check(r["ADDR1"], r["ADDR2"], r["ADDR3"], r["CITY"], r["ZIPCODE"]), axis=1)
    df = append_notes(df, notes)
    return df


def _is_valid_price(val) -> bool:
    try:
        float(str(val).strip())
        return True
    except ValueError:
        return False


def check_unfixable_product_prices(df: pd.DataFrame) -> None:
    """ProductPrice sometimes comes out corrupted (e.g. Excel stored it
    as an actual date). We can recompute it from TotalAmount using the
    country's VAT rate - but only for countries we have a rate for. Stop
    rather than guess for anything else."""
    bad = df[~df["ProductPrice"].apply(_is_valid_price)]
    if bad.empty:
        return
    unknown_country = bad[~bad["CountryName"].str.strip().str.lower().isin(VAT_RATES.keys())]
    if not unknown_country.empty:
        print("\n" + "=" * 70)
        print("STOPPED: the following rows have an invalid ProductPrice (e.g. it")
        print("looks like a date) and are in a country with no configured VAT rate,")
        print("so a net price can't be safely recalculated.")
        print("Add the country's standard VAT rate to VAT_RATES in the CONFIG")
        print("section, then re-run.")
        print("=" * 70)
        print(unknown_country[["IRID", "OrderNo", "CountryName", "ProductPrice", "TotalAmount"]]
              .drop_duplicates().to_string(index=False))
        sys.exit(1)

    bad_no_total = bad[~bad["TotalAmount"].apply(_is_valid_price)]
    if not bad_no_total.empty:
        print("\n" + "=" * 70)
        print("STOPPED: the following rows have an invalid ProductPrice AND an")
        print("invalid/unparseable TotalAmount, so a net price can't be recalculated.")
        print("=" * 70)
        print(bad_no_total[["IRID", "OrderNo", "CountryName", "ProductPrice", "TotalAmount"]]
              .drop_duplicates().to_string(index=False))
        sys.exit(1)


def fix_corrupted_product_price(df: pd.DataFrame) -> pd.DataFrame:
    """Where ProductPrice is invalid (fails to parse as a plain number -
    e.g. it came out as a date), recompute it as the net price before
    VAT: (TotalAmount - ShipFee) / (1 + country VAT rate). Valid prices
    are left untouched. Every row reaching this point is guaranteed to
    have a known VAT rate and a valid TotalAmount (checked beforehand)."""
    def fix_one(price, total, shipfee, country):
        if _is_valid_price(price):
            return str(price).strip(), ""
        vat = VAT_RATES[str(country).strip().lower()]
        total_f = float(str(total).strip())
        shipfee_f = float(str(shipfee).strip()) if str(shipfee).strip() else 0.0
        net = (total_f - shipfee_f) / (1 + vat)
        net_str = f"{net:.2f}"
        note = (f"ProductPrice edited: was '{price}' (invalid), recalculated as "
                f"{net_str} = net price before {int(vat * 100)}% VAT ({country}), "
                f"from TotalAmount {total_f:.2f}")
        return net_str, note

    results = df.apply(
        lambda r: fix_one(r["ProductPrice"], r["TotalAmount"], r["ShipFee"], r["CountryName"]),
        axis=1,
    )
    df["ProductPrice"] = [t[0] for t in results]
    df = append_notes(df, pd.Series([t[1] for t in results], index=df.index))
    return df


def check_bad_zips(df: pd.DataFrame) -> None:
    bad = df[df["ZIPCODE"].astype(str).str.strip().str.lower().isin(BAD_ZIPS)]
    if not bad.empty:
        print("\n" + "=" * 70)
        print("STOPPED: the following orders have missing/placeholder ZIP codes.")
        print("Supply the correct ZIP codes and either edit the source file")
        print("or hardcode them in this script's ZIP override section, then re-run.")
        print("=" * 70)
        for _, row in bad[["IRID", "IRNAME", "ADDR1", "CITY", "CountryName", "ZIPCODE"]].drop_duplicates().iterrows():
            print(f"  IRID {row.IRID} | {row.IRNAME} | {row.ADDR1}, {row.CITY}, {row.CountryName} "
                  f"| current ZIP: '{row.ZIPCODE}'")
        sys.exit(1)


def check_unmapped_substitution_products(df: pd.DataFrame) -> None:
    known_targets = set(PRODUCT_SUBSTITUTIONS.keys())
    # Flag any product name matching known out-of-stock patterns that ISN'T in our mapping
    watch_patterns = ["Pendant 3", "AirPure Zayn"]
    for pattern in watch_patterns:
        hits = df[df["ProdName"].str.contains(pattern, case=False, na=False)]
        if not hits.empty and pattern not in known_targets:
            print("\n" + "=" * 70)
            print(f"STOPPED: found product(s) matching '{pattern}' that need substitution,")
            print("but no substitution rule is configured in PRODUCT_SUBSTITUTIONS.")
            print("Add the correct replacement ProdCode/ProdName to the CONFIG section, then re-run.")
            print("=" * 70)
            print(hits[["IRID", "OrderNo", "ProdCode", "ProdName"]].to_string(index=False))
            sys.exit(1)


def check_combined_shipment_limit(df: pd.DataFrame) -> None:
    order_counts = df.groupby("IRID")["OrderNo"].nunique()
    over_limit = order_counts[order_counts > MAX_COMBINED_SHIPMENTS]
    if not over_limit.empty:
        print("\n" + "=" * 70)
        print(f"STOPPED: the following IR IDs have more than {MAX_COMBINED_SHIPMENTS} "
              f"distinct orders, exceeding the UPS combined-label limit.")
        print("These need manual review/splitting before proceeding.")
        print("=" * 70)
        print(over_limit.to_string())
        sys.exit(1)


def combine_orders_per_irid(df: pd.DataFrame) -> pd.DataFrame:
    """Where an IRID has 2-3 distinct orders, concatenate all OrderNo
    values (space separated) into the OrderNo column, and all ReceiptNo
    values (space separated) into the ReceiptNo column - kept separate,
    not merged together."""
    def combine(group):
        pairs = group[["OrderNo", "ReceiptNo"]].drop_duplicates()
        if len(pairs) > 1:
            orders = " ".join(pairs["OrderNo"])
            receipts = " ".join(pairs["ReceiptNo"])
            return pd.Series({"OrderNo": orders, "ReceiptNo": receipts})
        return None  # no change needed

    combined_map = df.groupby("IRID").apply(combine, include_groups=False)
    needs_combine = combined_map.dropna(how="all")
    if not needs_combine.empty:
        mask = df["IRID"].isin(needs_combine.index)
        df.loc[mask, "OrderNo"] = df.loc[mask, "IRID"].map(needs_combine["OrderNo"])
        df.loc[mask, "ReceiptNo"] = df.loc[mask, "IRID"].map(needs_combine["ReceiptNo"])
    return df


def apply_product_substitutions(df: pd.DataFrame) -> pd.DataFrame:
    for old_name, new_vals in PRODUCT_SUBSTITUTIONS.items():
        mask = df["ProdName"] == old_name
        if mask.any():
            df.loc[mask, "ProdName"] = new_vals["ProdName"]
            df.loc[mask, "ProdCode"] = new_vals["ProdCode"]
    return df


def fix_incomplete_home_numbers(df: pd.DataFrame) -> pd.DataFrame:
    """If ContactNoHome looks incomplete (too short to be a real number,
    e.g. just a country code, or blank), fall back to ContactNoMobile."""
    def is_incomplete(val: str) -> bool:
        digits = re.sub(r"\D", "", str(val))
        return len(digits) < 6  # shorter than any real subscriber number
    mask = df["ContactNoHome"].apply(is_incomplete)
    df.loc[mask, "ContactNoHome"] = df.loc[mask, "ContactNoMobile"]
    return df


def compute_packages(df: pd.DataFrame) -> pd.DataFrame:
    """Rule: 2 packages only if an IRID has >1 HomePure Zayn, OR has both
    HomePure Zayn AND HomePure Nova (Pi-Plus) together. Otherwise 1."""
    def pkg_for_group(group):
        zayn_count = group["ProdName"].str.contains(PACKAGE_TRIGGER_ZAYN, case=False, na=False).sum()
        has_nova = group["ProdName"].str.contains(PACKAGE_TRIGGER_NOVA, case=False, na=False).any()
        return 2 if (zayn_count > 1 or (zayn_count >= 1 and has_nova)) else 1

    pkg_map = df.groupby("IRID", group_keys=False).apply(pkg_for_group)
    df["Nr. of packages"] = df["IRID"].map(pkg_map)
    return df


def main():
    explicit = sys.argv[1] if len(sys.argv) > 1 else None
    src = find_input_file(explicit)
    print(f"Reading: {src}")

    df = load_raw(src)
    orig_len = len(df)
    df["Notes and Warnings"] = ""
    df = apply_zip_overrides(df)

    # --- Validation gates: stop rather than guess ---
    check_bad_zips(df)
    check_unmapped_substitution_products(df)
    check_combined_shipment_limit(df)
    check_unfixable_product_prices(df)

    # --- Transform ---
    df = df.sort_values(by="IRID", kind="stable").reset_index(drop=True)
    df = apply_product_substitutions(df)
    df = combine_orders_per_irid(df)
    df = fix_incomplete_home_numbers(df)
    df = compute_packages(df)
    df = pad_zip_codes(df)
    df = normalize_address_number_spacing(df)
    df = remove_duplicate_address_fields(df)
    df = split_long_addresses(df)
    df = flag_city_zip_in_address(df)  # run last, checks final address/zip values
    df = fix_corrupted_product_price(df)

    assert len(df) == orig_len, "Row count changed during processing - aborting, this should never happen."

    # Put Notes and Warnings last, after Nr. of packages (it was tracked
    # internally from the start so earlier steps could append to it).
    cols = [c for c in df.columns if c != "Notes and Warnings"] + ["Notes and Warnings"]
    df = df[cols]

    out_path = output_path_for(src)
    write_output(df, out_path)

    print(f"\nDone. {len(df)} rows written to:\n  {out_path}")
    print("Review the file, then upload it to SOMS manually as usual.")


if __name__ == "__main__":
    main()
